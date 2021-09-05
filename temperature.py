import xlrd
import xlwt
import paramiko
import os
import time
import re
import openpyxl
from openpyxl import Workbook

CSheet = xlrd.open_workbook(r'C:\Users\sbspf\Desktop\工控机程序更新\IPList.xlsx')
data = CSheet.sheets()[0]
nrow = data.nrows

FailedList = []

dict = {}

for i in range(1, nrow):
    ip = data.row_values(i)[0]

    try:
        transport = paramiko.Transport((ip, 22))
        transport.connect(username='root', password='dhERIS@2018*#')

        ssh = paramiko.SSHClient()
        # 自动添加策略，保存服务器的主机名和秘钥信息，不添加，不在本地hnow_hosts文件中的记录将无法连接
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh._transport = transport

        print("%s 连接成功" % (ip))

        # 查看硬盘温度
        stdin, stdout, stderror = ssh.exec_command("smartctl -x /dev/sda | grep Temperature_Celsius | awk '{print$8}'")
        temper = stdout.readlines()[0]

        dict[ip] = temper
        print(dict)

    except Exception as reason:
        print(reason)
        FailedList.append(ip)

# 创建表格
wb = Workbook()
ws = wb.create_sheet('sheet', 0)

i = 0
for ip in dict:  
    ws.cell(row=i+1,column=1).value = ip
    ws.cell(row=i+1,column=2).value = dict[ip]
    i += 1
    
    wb.save(r'C:\Users\sbspf\Desktop\temperature.xlsx')  #表格保存路径
    print('文件 导出成功！')

print("连接失败工控机：")
for j in FailedList:
    print(j)
