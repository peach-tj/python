import xlrd
import paramiko
import os
import datetime
from openpyxl import Workbook

CSheet = xlrd.open_workbook(r'C:\Users\Administrator\Desktop\Python\IPList.xlsx')
data = CSheet.sheets()[0]
nrow = data.nrows

FailedList = []


# 远程下载文件
def down_from_remote(sftp_obj, remote_dir_name, local_dir_name):
    if os.path.isdir(remote_dir_name):
        # 文件夹，不能直接下载，需要继续循环
        check_local_dir(local_dir_name)
        print('开始下载文件夹：' + remote_dir_name)
        for remote_file_name in sftp.listdir(remote_dir_name):
            sub_remote = os.path.join(remote_dir_name, remote_file_name)
            sub_remote = sub_remote.replace('\\', '/')
            sub_local = os.path.join(local_dir_name, remote_file_name)
            sub_local = sub_local.replace('\\', '/')
            down_from_remote(sftp_obj, sub_remote, sub_local)
    else:
        # 文件，直接下载
        print('开始下载文件：' + remote_dir_name)
        sftp_obj.get(remote_dir_name, local_dir_name)


# 本地文件夹是否存在，不存在则创建
def check_local_dir(local_dir_name):
    if not os.path.exists(local_dir_name):
        os.makedirs(local_dir_name)


# 获取文件内容
def check_file(local_dir_name, ws):
    with open(local_dir_name) as f:
        lines = f.readlines()      #读取全部内容 ，并以列表方式返回 
        m = 1 
        for i in range(1, len(lines)) :
            t2 = datetime.datetime.strptime(('20' + lines[i].replace('\n','').replace('_',' ')), "%Y-%m-%d %H:%M:%S")
            t1 = datetime.datetime.strptime(('20' + lines[i-1].replace('\n','').replace('_',' ')), "%Y-%m-%d %H:%M:%S")
            tt = (t2 - t1).seconds
            
            if tt > 180 :
                ws.cell(row=m+1, column=1).value = lines[i-1]
                ws.cell(row=m+2, column=1).value = lines[i]
                m = m+2


for i in range(1, nrow):
    ip = data.row_values(i)[0]

    fileName = ip.replace(".","_") + "_diskOK.txt"

    remotePath = "/mnt/soft/ERI-Server/diskOK.txt"
    localPath = "C:/Users/Administrator/Desktop/Python/DiskFile/%s"%(fileName)

    try:
        transport = paramiko.Transport((ip, 22))
        transport.connect(username='root', password='dhERIS@2018*#')

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())   # 自动添加策略，保存服务器的主机名和秘钥信息，不添加，不在本地hnow_hosts文件中的记录将无法连接
        ssh._transport = transport
        sftp = paramiko.SFTPClient.from_transport(transport)

        print("%s 连接成功" % (ip))
    
        # down_from_remote(sftp, remotePath, localPath)

        # 创建表格
        wb = Workbook()
        ws = wb.create_sheet('Sheet1', 0)

        check_file(localPath, ws)
        if ws.max_row > 1 :
            print('正在到处 %s 异常数据……'%(ip))
            wb.save('C:/Users/Administrator/Desktop/Python/DiskFile/%s.xlsx'%(ip.replace('.','_')))

    except Exception as reason:
        print(reason)
        FailedList.append(ip)


print('连接异常工控机:')
for j in FailedList :
    print(j)

