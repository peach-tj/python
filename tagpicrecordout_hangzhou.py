import openpyxl
from openpyxl import Workbook
import cx_Oracle
import xlrd
import xlwt
import paramiko


Sheet = xlrd.open_workbook(r'C:\Users\Administrator\Desktop\Python\工控机程序更新\IPList.xlsx')
data = Sheet.sheets()[0]
nrow = data.nrows

rowCount = 0
TimeoutList = []

# 创建表格
wb = Workbook()
ws = wb.create_sheet('Sheet1', 0)

for i in range(1, nrow):
    ip = data.row_values(i)[0]
    conInfo = "ERIS/ERIS@%s:1521/ERIS"%(ip)

    sql="select '17:00-21:00'时间段,max(c.name) 区域,max(b.name) 路口名,round(avg(to_number(a.CREATE_DATE-a.pic_cap_date)*24*60*60-0.5),2)入库时间,count(*)总数 from t_tag_pic_record  a inner join t_organization b on a.ORGANIZA_ID=b.id inner join t_organization c on b.parent_id=c.id where to_char(a.pic_cap_date,'yyyymmdd hh24:mi:ss')between '20210810 17:00:00' and   '20210810 21:00:00' and a.RECORD_TYPE in('1','2')"
    
    try:
        # 执行sql查询
        con = cx_Oracle.connect(conInfo)
        cur = con.cursor()
        cur.execute(sql)
        results= cur.fetchall()

        # 获取表字段值
        db_title = [i[0] for i in cur.description]
        for i, description in enumerate(db_title):
            ws.cell(row=1, column=2+i).value = description
        ws.cell(row=1, column=1).value = "IP"

        ws.cell(row=rowCount+2, column=1).value = ip
        ws.cell(row=rowCount+2, column=2).value = results[0][0]
        ws.cell(row=rowCount+2, column=3).value = results[0][1]
        ws.cell(row=rowCount+2, column=4).value = results[0][2]
        ws.cell(row=rowCount+2, column=5).value = results[0][3]
        ws.cell(row=rowCount+2, column=6).value = results[0][4]

        rowCount = rowCount + 1

        print("正在存储 %s 巡检数据……"%(ip))
        wb.save(r'C:\Users\Administrator\Desktop\xunjian.xlsx')      

    except Exception as err:
        print("%s 连接超时： "%(ip), err)
        TimeoutList.append(ip)

if len(TimeoutList) > 0 :
    print("连接超时的工控机有：")
    for j in TimeoutList :
        print(j)