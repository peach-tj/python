import openpyxl
from openpyxl import Workbook
import cx_Oracle
import datetime
import time
import xlrd
import datetime
from time import strftime


con = cx_Oracle.connect('ERIS/ERISRP@33.83.247.42:1521/ERIS')       # 连接数据库
cur = con.cursor() # 获取游标

start = time.clock()

CSheet = xlrd.open_workbook(r'C:\Users\sbspf\Desktop\工控机程序更新\tag_pic_record_count\hphm.xlsx',encoding_override='utf-8')   # 需要查找的号牌表，将此处修改为表格所在的路径及文件名
data = CSheet.sheets()[0]
nrows = data.nrows


cur_time = strftime("%Y-%m-%d %H:%M:%S",time.localtime())
start_time = "2021-05-21 00:00:00"
end_time = "2021-05-21 23:59:59"

cameras = []
rfids = []
matches =[]

def init_data() :
    start_time = "2021-05-21 00:00:00"
    end_time = "2021-05-21 23:59:59"
    cameras = []
    rfids = []
    matches =[]
    return start_time, end_time, cameras, rfids, matches


# 创建表格
wb = Workbook()
ws = wb.create_sheet('Sheet1',0)

ws.cell(row=1,column=1).value = '号牌号码'
ws.cell(row=1,column=2).value = '视频总数'
ws.cell(row=1,column=3).value = '射频总数'
ws.cell(row=1,column=4).value = '融合总数'


for i in range(1,nrows):
    result = data.row_values(i)

    hphm = result[1]        # 获取号牌号码
    ws.cell(row=i+1,column=1).value = hphm  # 将号牌号码存储在新的表格
    print("正在统计 %s 过车数…"%(hphm))

    if type(hphm) == float :
        hphm = '%d'%hphm
    else :
        hphm = ''.join(filter(lambda c:ord(c)<256, hphm))   # 去除汉字


    while(start_time <= cur_time):

        sql = "select count(*) from t_tag_pic_record where (hphm = '%s' or hphm = '杭州%s') and record_type = 1 and (cap_date between to_date('%s','yyyy-mm-dd hh24:mi:ss') and to_date('%s','yyyy-mm-dd hh24:mi:ss'))"%(hphm, hphm, start_time, end_time)       # 视频过车数量
        rfidsql = "select count(*) from t_tag_pic_record where (hphm = '%s' or hphm = '杭州%s') and record_type = 0 and (cap_date between to_date('%s','yyyy-mm-dd hh24:mi:ss') and to_date('%s','yyyy-mm-dd hh24:mi:ss'))"%(hphm, hphm, start_time, end_time)   # 射频过车数量
        matchsql = "select count(*) from t_tag_pic_record where (hphm = '%s' or hphm = '杭州%s') and record_type = 2 and (cap_date between to_date('%s','yyyy-mm-dd hh24:mi:ss') and to_date('%s','yyyy-mm-dd hh24:mi:ss'))"%(hphm, hphm, start_time, end_time)  # 融合过车数量

        cur.execute(sql)
        camera = cur.fetchone()
        cameras.append(camera[0])

        cur.execute(rfidsql)
        rfid = cur.fetchone()
        rfids.append(rfid[0])

        cur.execute(matchsql)
        matche = cur.fetchone()
        matches.append(matche[0])

        start_time = (datetime.datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S") + datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
        end_time = (datetime.datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S") + datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
    
    ws.cell(row=i+1,column=2).value = sum(cameras)    # 视频
    ws.cell(row=i+1,column=3).value = sum(rfids)     # 射频
    ws.cell(row=i+1,column=4).value = sum(matches)     # 融合

    start_time, end_time, cameras, rfids, matches = init_data()

    wb.save(filename = 'match_record.xlsx')     # 保存文件，路径默认为脚本所在当前目录


cur.close() # 关闭游标
con.close() # 关闭链接


end = time.clock()
print('运行时间：%s' %(end - start))
