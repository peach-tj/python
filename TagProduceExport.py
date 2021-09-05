import openpyxl
from openpyxl import Workbook
import cx_Oracle

con = cx_Oracle.connect('ERIS/ERIS@172.17.2.77:1521/ERIS')       # 链接数据库
cur = con.cursor()  # 获取游标

sql1 = "select plate_num,batch_num,bicycle_owner_type,express_company,qr_code,status,rfid_num,express_num,create_date,feedback_date,upload_date,area_street_community,site_num,branch_name from t_pingan_plate_info where plate_num like 'B%' and create_date < to_date('2020-11-15 23:59:59','yyyy-mm-dd hh24:mi:ss') order by create_date desc"

sql2 = "select * from t_pingan_plate_info where bicycle_owner_type = 4"

sql3 = "select id,hphm,TID from t_tag_produce_manager where FPDH = '杭州'"

sql = "select * from t_pingan_plate_info where bicycle_owner_type = 4 and create_date <= to_date('2021-07-20 12:00:00','yyyy-mm-dd hh24:mi:ss') order by id desc"

# 执行sql查询
cur.execute(sql)
results = cur.fetchall()

# 获取行和列
rows = len(results)
if len(results):
    cols = len(results[0])  # 判断list是否溢出

# 创建表格
wb = Workbook()
ws = wb.create_sheet('Sheet1', 0)

# 获取表字段值
db_title = [i[0] for i in cur.description]
for i, description in enumerate(db_title):
    ws.cell(row=1, column=1+i).value = description


# 循环保存查询数据
for m in range(rows):
    for n in range(cols):
        ws.cell(row=m+2, column=n+1).value = results[m][n]
    print(results[m])

    wb.save('pingan_plate.xlsx')


cur.close()  # 关闭游标
con.close()  # 关闭链接

