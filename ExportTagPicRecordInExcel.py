import openpyxl
from openpyxl import Workbook
import cx_Oracle
import datetime
import time
import xlrd
import datetime
from time import strftime

con = cx_Oracle.connect('ERIS/ERIS@172.17.2.35:1521/ERIS')       # 连接数据库
cur = con.cursor() # 获取游标

start = time.clock()

CSheet = xlrd.open_workbook(r'C:\Users\Administrator\Desktop\PythonTest\hphm.xlsx',encoding_override='utf-8')   # 需要查找的号牌表，将此处修改为表格所在的路径及文件名
data = CSheet.sheets()[0]
nrows = data.nrows

cur_time = strftime("%Y-%m-%d %H:%M:%S",time.localtime())
start_time = "2021-06-26 00:00:00"
end_time = "2021-06-26 23:59:59"

hphmList1 = []
hphmList2 = []
recordData = []

for i in range(1,nrows):
    hphm = data.row_values(i)[0]
    hphmList1.append(hphm)

    if type(hphm) == float :
        hphm = '%d'%hphm
    else :
        hphm = ''.join(filter(lambda c:ord(c)<256, hphm))   # 去除汉字
    hphmList2.append(hphm)


while(start_time <= cur_time):

    sql = ("select hphm, reader_name, dev_name, cap_date, record_type from t_tag_pic_record where (hphm in %s or hphm in %s) and record_type = 1 and (cap_date between to_date('%s','yyyy-mm-dd hh24:mi:ss') and to_date('%s','yyyy-mm-dd hh24:mi:ss'))"%(hphmList2, hphmList1, start_time, end_time)).replace("[",'(').replace("]",')')    # 视频过车数量
    rfidsql = ("select hphm, reader_name, dev_name, cap_date, record_type from t_tag_pic_record where (hphm in %s or hphm in %s) and record_type = 0 and (cap_date between to_date('%s','yyyy-mm-dd hh24:mi:ss') and to_date('%s','yyyy-mm-dd hh24:mi:ss'))"%(hphmList2, hphmList1, start_time, end_time)).replace("[",'(').replace("]",')')   # 射频过车数量
    matchsql = ("select hphm, reader_name, dev_name, cap_date, record_type from t_tag_pic_record where (hphm in %s or hphm in %s) and record_type = 2 and (cap_date between to_date('%s','yyyy-mm-dd hh24:mi:ss') and to_date('%s','yyyy-mm-dd hh24:mi:ss'))"%(hphmList2, hphmList1, start_time, end_time)).replace("[",'(').replace("]",')')  # 融合过车数量

    cur.execute(sql)
    result1 = cur.fetchall()
    for i in range(0, len(result1)):
        recordData.append(result1[i])

    cur.execute(rfidsql)
    result2 = cur.fetchall()
    for i in range(0, len(result2)):
        recordData.append(result2[i])

    cur.execute(matchsql)
    result3 = cur.fetchall()
    for i in range(0, len(result3)):
        recordData.append(result3[i])

    start_time = (datetime.datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S") + datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
    end_time = (datetime.datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S") + datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")


# 创建表格
wb = Workbook()
ws = wb.create_sheet('Sheet1', 0)


# 获取表字段值
db_title = [i[0] for i in cur.description]
for i,description in enumerate(db_title):
    ws.cell(row=1,column=1+i).value = description

for k in range(len(recordData)):
    print(recordData[k])
    for h in range(len(recordData[k])):
        ws.cell(row=k+2, column=h+1).value = recordData[k][h]

wb.save('tag_pic_record.xlsx')
print('文件导出成功！')
